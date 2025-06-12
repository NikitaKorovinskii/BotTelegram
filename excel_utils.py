import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

FILE_NAME = "workout_data.xlsx"

def save_to_excel(username, data):
    new_row = {
        "Дата": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Упражнение": data["exercise"],
        "Повторения": int(data["reps"]),
        "Вес": float(data["weight"])
    }

    try:
        book = load_workbook(FILE_NAME)
    except FileNotFoundError:
        df = pd.DataFrame([new_row])
        with pd.ExcelWriter(FILE_NAME, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=username, index=False)
        return

    if username in book.sheetnames:
        df = pd.read_excel(FILE_NAME, sheet_name=username)
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    else:
        df = pd.DataFrame([new_row])

    with pd.ExcelWriter(FILE_NAME, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=username, index=False)

def get_user_exercises(username):
    try:
        df = pd.read_excel(FILE_NAME, sheet_name=username)
        return df["Упражнение"].dropna().unique().tolist()
    except Exception:
        return []
